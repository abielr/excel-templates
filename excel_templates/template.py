from copy import copy
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

class ExcelTemplate(object):
    def __init__(self, filename):
        """
        :param filename: name of Excel template file
        """

        self.wb = load_workbook(filename)
        self.wbv = load_workbook(filename, data_only=True)
        self.original_dimensions = dict((ws.title, (ws.max_row, ws.max_column)) for ws in self.wb.worksheets)
        self.updated_dimensions = dict((ws.title, (ws.max_row, ws.max_column)) for ws in self.wb.worksheets)
        self.grid_dimensions = dict((ws.title, (1,1)) for ws in self.wb.worksheets)
        self.tiled_sheets = set()

    def copy_worksheet(self, source_name, target_name):
        """
        Copy an Excel worksheet

        :param source_name: source worksheet name
        :param target_name: new worksheet name
        :return:
        """

        target = self.wb.copy_worksheet(self.wb[source_name])
        target.title = target_name
        target.freeze_panes = self.wb[source_name].freeze_panes

        target = self.wbv.copy_worksheet(self.wbv[source_name])
        target.title = target_name

        self.original_dimensions[target_name] = self.original_dimensions[source_name]
        self.updated_dimensions[target_name] = self.updated_dimensions[source_name]
        self.grid_dimensions[target_name] = self.grid_dimensions[source_name]

    def tile(self, sheetname, rows, columns, row_spacing=1, col_spacing=1):
        """
        Repeat the contents of an Excel worksheet, tiling it into a grid

        :param sheetname: name of Excel worksheet
        :param rows: number of times to repeat down the rows
        :param columns: number of times to repeat across the columns
        :param row_spacing: number of rows to insert between each repeated template
        :param col_spacing: number of columns to insert between each repeated template
        :return:
        """

        if sheetname in self.tiled_sheets:
            raise Exception("Can only expand the sheet '{sheet}' once".format(sheet=sheetname))
        self.tiled_sheets.add(sheetname)

        ws = self.wb[sheetname]
        rng = ws[ws.dimensions]

        row_offset = ws.max_row + row_spacing
        col_offset = ws.max_column + col_spacing
        self.updated_dimensions[sheetname] = (row_offset, col_offset)
        self.grid_dimensions[sheetname] = (rows, columns)

        merged_cells = [(r.min_row, r.min_col, r.max_row, r.max_col) for r in ws.merged_cells.ranges]

        for j in range(columns):
            for i in range(rows):
                if i == 0 and j == 0: continue
                for row in rng:
                    for source_cell in row:
                        target_cell = ws.cell(source_cell.row + row_offset*i, source_cell.column + col_offset*j)

                        if not isinstance(source_cell.value, int):
                            target_cell._value = Translator(source_cell._value, origin=source_cell.coordinate). \
                                translate_formula(target_cell.coordinate)
                        else:
                            target_cell.value = source_cell.value
                        target_cell.data_type = source_cell.data_type

                        if source_cell.has_style:
                            target_cell._style = copy(source_cell._style)

                        if source_cell.hyperlink:
                            target_cell._hyperlink = copy(source_cell.hyperlink)

                        if source_cell.comment:
                            target_cell.comment = copy(source_cell.comment)

                for min_row, min_col, max_row, max_col in merged_cells:
                    ws.merge_cells(start_row=min_row+row_offset*i, start_column=min_col+col_offset*j,
                                   end_row=max_row+row_offset*i, end_column=max_col+col_offset*j)

        for i in range(1,rows):
            for ii in range(len(rng)):
                ws.row_dimensions[ii+1+row_offset*i].height = ws.row_dimensions[ii+1].height

        for j in range(1,columns):
            for jj in range(len(rng[0])):
                ws.column_dimensions[get_column_letter(jj+col_offset*j+1)].width = \
                    ws.column_dimensions[get_column_letter(jj + 1)].width

    def fill(self, sheetname, data, grid_row=1, grid_col=1, prefix='', fillna=None):
        """
        Fill keys in a worksheet with corresponding values from Python dictionary

        :param sheetname: name of Excel worksheet
        :param data: dictionary object with values that will be used to fill in template
        :param grid_row: if you have tiled the template, the row number of the tiled template
        :param grid_col: if you have tiled the template, the column number of the tiled template
        :param prefix: only fill in cells that begin with given prefix
        :param fillna: if key is missing in data, fill with fillna. If None than ignore
        :return:
        """

        if grid_row > self.grid_dimensions[sheetname][0] or grid_col > self.grid_dimensions[sheetname][1]:
            raise Exception("Invalid cell position (%d, %d)" % (grid_row, grid_col))
        ws = self.wb[sheetname]
        row_offset, col_offset = self.updated_dimensions.get(sheetname, self.original_dimensions[sheetname])
        orow_offset, ocol_offset = self.original_dimensions[sheetname]

        for irow in range(1, orow_offset+1):
            for icol in range(1, ocol_offset+1):
                newrow = irow + (grid_row-1)*row_offset
                newcol = icol + (grid_col-1)*col_offset
                key = str(self.wbv[sheetname].cell(irow, icol).value)
                if key.startswith(prefix):
                    if key[len(prefix):] in data:
                        ws.cell(newrow, newcol).value = data[key[len(prefix):]]
                    elif not fillna is None:
                        ws.cell(newrow, newcol).value = fillna

    def save(self, filename):
        self.wb.save(filename)

def make_dict(df, keys, value, sep):
    """
    Turn a DataFrame into a dictionary by joining column values
    :param df: input DataFrame
    :param keys: list of column names to join together
    :param value: column containing values
    :param sep: string to join on
    :return: a dictionary
    """
    d = {}
    for key, value in zip(df[keys].astype(str).apply(lambda row: sep.join(row.values), axis=1), df[value]):
        d[key] = value
    return d
